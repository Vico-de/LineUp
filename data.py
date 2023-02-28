import random

import openpyxl

wb = openpyxl.load_workbook('LineTest.xlsx')


def artists():
    """
    Renvoie un dictionnaire avec les artistes.
    """
    sheet = wb['Artistes']

    artist_cards = {}
    for row in sheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if None in values:
            continue
        artist_name, artist_label, artist_cost, artist_stars, artist_scene, artist_style, artist_gender = values
        artist_cards[artist_name] = {
            'label': artist_label,
            'cost': artist_cost,
            'stars': artist_stars,
            'style': artist_scene,
            'scene': artist_style,
            'genre': artist_gender,
        }

    return artist_cards


def scenes():
    """
    Renvoie un dictionnaire avec les scènes.
    """
    sheet = wb['Scenes']

    # Dico des scènes
    scene_cards = {}
    for row in sheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if None in values:
            continue
        scene_name, scene_number, scene_cost, scene_type, scene_stars, wrong_style = values
        scene_cards[scene_name] = {'scene': scene_number, 'cost': scene_cost, 'type': scene_type,
                                   'stars': scene_stars, "wrong style": wrong_style}

    return scene_cards


def events():
    """
    Renvoie un dictionnaire avec les cartes événements.
    """
    sheet = wb['Events']

    events_cards = {}
    for row in sheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if None in values:
            continue
        name, event, effect, fame, style = values
        events_cards[name] = {
            'event': event,
            'effect': effect,
            'fame': fame,
            'style': style.split(),
        }

    return events_cards


def random_list(dictionary):
    stack = list(dictionary.keys())
    random.shuffle(stack)
    return stack
