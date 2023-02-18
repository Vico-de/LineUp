import random

import openpyxl

wb = openpyxl.load_workbook('LineTest.xlsx')
# Dico des artistes
sheet = wb['Artistes']

artist_card = {}
for row in sheet.iter_rows(min_row=2):
    values = [cell.value for cell in row]
    if None in values:
        continue
    artist_name, artist_label, artist_cost, artist_stars, artist_scene, artist_style, artist_gender = values
    artist_card[artist_name] = {'Label': artist_label, 'Prix': artist_cost, 'Etoiles': artist_stars,
                                'Scene': artist_scene, 'Style': artist_style, 'Genre': artist_gender}

sheet = wb['Scenes']

# Dico des scènes
scene_card = {}
for row in sheet.iter_rows(min_row=2):
    values = [cell.value for cell in row]
    if None in values:
        continue
    scene_name, scene_cost, scene_type, scene_stars = values
    scene_card[scene_name] = {'Prix': scene_cost, 'Type': scene_type, 'Etoiles': scene_stars}

# Définition du nombre de joueurs, de leur budget et de leur dictionnaire
num_players = int(input("Entrez le nombre de joueurs : "))
players = []
for i in range(num_players):
    player_dict = {'name': f"Joueur {i + 1}", 'budget': 1000000, 'choice': "ok",
                   'inventory': {'artists': [], 'scenes': []}}
    players.append(player_dict)

## CODE TEMPORAIRE POUR TEST PLUS RAPIDEMENT
artist_card_stack = list(artist_card.keys())
scene_card_stack = list(scene_card.keys())

# Mélanger les cartes artistes et les répartir entre les joueurs
random.shuffle(artist_card_stack)
for i in range(len(artist_card_stack)):
    player = players[i % num_players]
    player['inventory']['artists'].append(artist_card_stack[i])

# Mélanger les cartes scènes et les répartir entre les joueurs
random.shuffle(scene_card_stack)
for i in range(len(scene_card_stack)):
    player = players[i % num_players]
    player['inventory']['scenes'].append(scene_card_stack[i])

# # Afficher les decks de chaque joueur
# for player in players:
#     print(player['name'])
#     print("Cartes artistes :", player['inventory']['artists'])
#     print("Cartes scènes :", player['inventory']['scenes'])
#     print("\n")


#CODE TEMPORAIRE POUR DISPATCH LES ARTISTES
def assign_artists_to_scenes(players):
    for player in players:
        # Pour chaque joueur, on itère sur tous les artistes qu'il possède
        for artist_name in player['inventory']['artists']:
            # On récupère le dictionnaire de l'artiste
            artist = artist_card[artist_name]
            # On choisit une scène aléatoire dans la liste des scènes du joueur
            scene_name = random.choice(player['inventory']['scenes'])
            # On récupère le dictionnaire de la scène
            scene = scene_card[scene_name]
            # On ajoute l'artiste à la liste des artistes de la scène
            if 'artists' not in scene:
                scene['artists'] = {}
            if artist_name not in scene['artists']:
                scene['artists'][artist_name] = {
                    'Label': artist['Label'],
                    'Prix': artist['Prix'],
                    'Etoiles': artist['Etoiles'],
                    'Style': artist['Style'],
                    'Genre': artist['Genre']
                }

    # On affiche l'inventaire mis à jour de chaque joueur
    display_inventory(players)

assign_artists_to_scenes(player_list, artist_card, scene_card)



# ---------------------------------------------------------------------
# # Liste des cartes scènes disponibles
# scene_card_stack = list(scene_card.keys())
#
# # Boucle pour l'achat des cartes scènes
# while scene_card_stack:
#     for player in players:
#         print(f"Achetez une carte scène ou passez votre tour, {player['name']} (budget : {player['budget']})")
#         print("Scènes disponibles :", scene_card_stack)
#         choice = input("Entrez le nom de la carte ou 'passe' pour passer votre tour : ")
#
#         # Si le joueur passe son tour
#         if choice.lower() == "passe":
#             player['choice'] = "passe"
#             continue
#
#         # Si le choix est une carte disponible
#         if choice in scene_card_stack:
#             card_cost = scene_card[choice]['Prix']
#
#             # Vérification si le joueur a suffisamment d'argent pour acheter la carte
#             if player['budget'] >= card_cost:
#                 player['budget'] -= card_cost  # Déduction du coût de la carte du budget du joueur
#                 player['inventory']['scenes'].append(choice)  # Ajout de la carte achetée dans le dictionnaire du joueur
#                 scene_card_stack.remove(choice)  # Retrait de la carte achetée de la liste des cartes disponibles
#                 print(f"{player['name']} a acheté la carte {choice} pour {card_cost} écocups.")
#                 print(f"Il te reste {player['budget']} écocup !")
#             else:
#                 print(f"{player['name']}, vous n'avez pas assez d'argent pour acheter cette carte.")
#
#     # Vérifier si tous les joueurs ont passé leur tour
#     if all(choice.lower() == "passe" for choice in (p['choice'] for p in players)):
#         break


# --------------------------------------------------------------------
# # Mélange des cartes
# artist_stack = list(artist_card.keys())
# random.shuffle(artist_stack)
#
# # Retirer 10 cartes random
# num_to_remove = round(0.2 * len(artist_stack))
# removed_random = random.sample(artist_stack, num_to_remove)
# remaining_stack = [artist for artist in artist_stack if artist not in removed_random]
#
# print(removed_random)
# print(remaining_stack)
#
# artist_discard = []
# # Simuler le déroulement d'une partie
# removed_card = []
# while len(artist_stack) > 0:
#     artist_name = artist_stack.pop()
#     print("\nLa carte retournée est :", artist_name)
#     artist_label, _, artist_stars, artist_style, artist_gender = artist_card[artist_name]
#     print("Le coût de cette carte est :", artist_card[artist_name]["Prix"])
#
#     # Boucle pour permettre aux joueurs de surenchérir ou de passer leur tour
#     highest_bidder = None
#     artist_cost = artist_card[artist_name]["Prix"]
#     last_bid = artist_cost
#     for i, player in enumerate(players):
#         player_budget = player['budget']
#         last_bid = 0  # initialisation à 0
#         if player_budget < last_bid:
#             print(f"{player} ne peut pas surenchérir car il n'a plus assez d'argent.")
#             continue
#         while True:
#             if i == 0:
#                 bid = input(
#                     f"{player}, proposez un montant égal ou supérieur à {last_bid} (ou tapez 'passer' pour passer votre tour) : ")
#                 if bid == 'passer':
#                     print(f"{player} passe son tour.")
#                     break
#
#                 try:
#                     bid = int(bid)
#                     if bid < artist_cost:
#                         print("Votre enchère doit être supérieure ou égale au prix initial.")
#                     elif bid > player_budget:
#                         print("Vous n'avez pas assez d'argent pour proposer ce montant.")
#                     elif bid % 1000 != 0:
#                         print("Votre enchère doit être un multiple de 1000.")
#                     else:
#                         last_bid = bid
#                         highest_bidder = player
#                         artist_cost = bid
#                         print(f"{player} propose un montant de {bid}.")
#                         break
#                 except ValueError:
#                     print("Vous devez entrer un nombre entier ou 'passer'.")
#             else:
#                 bid = input(
#                     f"{player}, proposez un montant supérieur à {last_bid} (ou tapez 'passer' pour passer votre tour) : ")
#                 if bid == 'passer':
#                     print(f"{player} passe son tour.")
#                     break
#
#                 try:
#                     bid = int(bid)
#                     if bid <= last_bid:
#                         print("Votre enchère doit être supérieure au montant précédent.")
#                     elif bid > player_budget:
#                         print("Vous n'avez pas assez d'argent pour proposer ce montant.")
#                     elif bid % 1000 != 0:
#                         print("Votre enchère doit être un multiple de 1000.")
#                     else:
#                         last_bid = bid
#                         highest_bidder = player
#                         print(f"{player} propose un montant de {bid}.")
#                         break
#                 except ValueError:
#                     print("Vous devez entrer un nombre entier ou 'passer'.")
#
#     # Vendre la carte au joueur ayant fait la plus haute enchère
#     if highest_bidder is not None:
#         highest_bidder['budget'] -= last_bid
#         print(f"{highest_bidder['name']} remporte la carte pour un montant de {last_bid}.")
#         # ajout de la carte au dictionnaire de l'inventaire du joueur
#         highest_bidder['inventory']['artists'].append(artist_name)
#
#         removed_card.append(artist_name)
#     else:
#         print("La carte est retirée de la partie.")
#
#     # Affichage des budgets restants
#     for player in players:
#         print(f"{player['name']} a maintenant {player['budget']} euros.")


# Soucis ici

# CODE TEMPORAIRE POUR DISPATCH LES ARTISTES DANS LES SCENES DE FACON ALEATOIRE








# # Dispatch des artistes dans les scènes
# for player in players:
#     inventory = player['inventory']
#     scenes = inventory['scenes']
#     artists = inventory['artists']
#     print(f"{player['name']}, voici les scènes disponibles dans votre deck :")
#     for scene_name, scene in scene_card.items():
#         if scene_name in scenes:
#             print(f"Scène : {scene_name}")
#     while artists:
#         artist = artists[0]
#         print(f"{player['name']}, voici les artistes disponibles dans votre deck : {artists}")
#         scene_choice = input(f"Dans quelle scène voulez-vous placer {artist} ? ")
#         if scene_choice in scenes:
#             scene = scene_card[scene_choice]
#             # Ajouter l'artiste à la scène
#             scene.setdefault('Artists', []).append(artist)
#             # Retirer l'artiste du deck du joueur
#             inventory['artists'].remove(artist)
#             artists = inventory['artists']
#         else:
#             print("Scène invalide, veuillez choisir une scène présente dans votre deck.")
#     print("Tous les artistes ont été répartis dans les scènes.")


# Comptage des points


# sheet = wb['Combo']
# # Créer un dictionnaire vide pour stocker les points gagnés par chaque joueur
# scores = {}
#
# # Parcourir chaque scène dans le deck de chaque joueur
# for player in players:
#     inventory = player['inventory']
#     scenes = inventory['scenes']
#     artists = inventory['artists']
#     player_name = player['name']
#     total_score = 0
#
#     for scene_name in scenes:
#         scene = scene_card[scene_name]
#         scene_artists = scene.get('Artists', [])
#         artist_labels = [artist_card[artist['name']]['Label'] for artist in scene_artists]
#
#         # Vérifier si les artistes de la scène satisfont une ou plusieurs conditions
#         for row in sheet.iter_rows(min_row=2):
#             condition1, condition2, points = [cell.value for cell in row]
#             if condition1 in artist_labels and condition2 in artist_labels:
#                 total_score += points
#
#     scores[player_name] = total_score
#     print(f"{player_name} a obtenu {total_score} points.")
